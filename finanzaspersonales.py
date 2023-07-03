import mysql.connector
from getpass import getpass
import openpyxl

# Conexión a la base de datos
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="azulazul34",
    database="gestion_finanzas"
)
mycursor = mydb.cursor()

# Función para validar el nombre de usuario
def validar_usuario(usuario):
    if not usuario.isalpha():
        print("El usuario solo debe contener letras.")
        return False
    return True

# Función para validar la contraseña
def validar_contrasena(contrasena):
    if len(contrasena) < 5:
        print("La contraseña debe tener al menos 5 caracteres.")
        return False
    if any(str(i) + str(i+1) in contrasena for i in range(9)):
        print("La contraseña no debe contener números consecutivos.")
        return False
    return True

# Función para iniciar sesión o crear usuario
def iniciar_sesion():
    usuario = input("Ingrese su nombre de usuario: ")
    contrasena = getpass("Ingrese su contraseña: ")

    # Verificar si el usuario existe
    mycursor.execute("SELECT id FROM Usuarios WHERE nombre = %s AND password = %s", (usuario, contrasena))
    result = mycursor.fetchone()

    if result:
        print("Ingreso exitoso!")
        return result[0]  # Devolver el id del usuario
    else:
        print("El usuario y/o contraseña son incorrectos.")

        # Preguntar si desea crear un nuevo usuario
        crear_usuario = input("¿Desea crear un nuevo usuario? (S/N): ")
        if crear_usuario.upper() == "S":
            while True:
                if validar_usuario(usuario):
                    contrasena = getpass("Ingrese su contraseña (debe tener al menos 5 caracteres y no contener números consecutivos): ")
                    if validar_contrasena(contrasena):
                        # Insertar nuevo usuario en la base de datos
                        mycursor.execute("INSERT INTO Usuarios (nombre, password) VALUES (%s, %s)", (usuario, contrasena))
                        mydb.commit()
                        print("Usuario creado exitosamente.")
                        return mycursor.lastrowid  # Devolver el id del usuario
                usuario = input("Ingrese su nombre de usuario: ")

# Función para mostrar las transacciones del usuario
def mostrar_transacciones(usuario_id):
    # Obtener las transacciones del usuario de la base de datos
    mycursor.execute("SELECT * FROM Transacciones WHERE usuario_id = %s", (usuario_id,))
    transacciones = mycursor.fetchall()

    if transacciones:
        print("Transacciones:")
        for transaccion in transacciones:
            print(transaccion)
    else:
        print("No hay transacciones para mostrar.")

# Función para agregar una transacción
def agregar_transaccion(usuario_id):
    tipo = input("Tipo (1: Ingreso, 2: Gasto): ")
    descripcion = input("Descripción: ")
    monto = float(input("Monto: "))
    fecha = input("Fecha (YYYY-MM-DD): ")

    # Insertar la nueva transacción en la base de datos
    mycursor.execute("INSERT INTO Transacciones (descripcion, tipo, monto, fecha, usuario_id) VALUES (%s, %s, %s, %s, %s)",
                     (descripcion, tipo, monto, fecha, usuario_id))
    mydb.commit()
    print("La transacción ha sido guardada con éxito.")

# Función para generar el reporte en Excel
def generar_reporte(usuario_id):
    # Obtener las transacciones del usuario de la base de datos
    mycursor.execute("SELECT * FROM Transacciones WHERE usuario_id = %s", (usuario_id,))
    transacciones = mycursor.fetchall()

    if transacciones:
        # Crear un nuevo archivo de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Escribir las transacciones en el archivo de Excel
        for i, transaccion in enumerate(transacciones, start=1):
            sheet.cell(row=i, column=1, value=transaccion[3].upper())
            sheet.cell(row=i, column=2, value=transaccion[2])
            sheet.cell(row=i, column=3, value=transaccion[4])
            sheet.cell(row=i, column=4, value=transaccion[5])
            if transaccion[2] == "Gasto":
                sheet.cell(row=i, column=1).fill = openpyxl.styles.PatternFill(start_color="0000FF",
                                                                              end_color="0000FF",
                                                                              fill_type="solid")

        # Calcular el balance
        ingresos = sum(transaccion[4] for transaccion in transacciones if transaccion[2] == "Ingreso")
        gastos = sum(transaccion[4] for transaccion in transacciones if transaccion[2] == "Gasto")
        balance = ingresos - gastos

        # Escribir el balance en el archivo de Excel
        sheet.cell(row=len(transacciones) + 2, column=1, value="Balance")
        sheet.cell(row=len(transacciones) + 2, column=2, value=balance)
        sheet.cell(row=len(transacciones) + 2, column=1).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=len(transacciones) + 2, column=2).font = openpyxl.styles.Font(bold=True)

        # Escribir el mensaje de autoría
        sheet.cell(row=len(transacciones) + 4, column=1, value="Hecho por Isaac S. :)")

        # Guardar el archivo de Excel
        workbook.save("reporte.xlsx")

        print("El reporte se ha generado correctamente.")
    else:
        print("No hay transacciones para generar el reporte.")

# Programa principal
def main():
    usuario_id = iniciar_sesion()

    if usuario_id:
        while True:
            print("\nMenu:")
            print("1. Mostrar transacciones")
            print("2. Agregar transacción")
            print("3. Generar reporte en Excel")
            print("4. Cerrar")

            opcion = input("Ingrese la opción deseada: ")

            if opcion == "1":
                mostrar_transacciones(usuario_id)
            elif opcion == "2":
                agregar_transaccion(usuario_id)
            elif opcion == "3":
                generar_reporte(usuario_id)
            elif opcion == "4":
                print("Gracias por utilizarme, ten un buen día.")
                break
            else:
                print("Opción inválida.")

# Ejecutar el programa principal
main()
